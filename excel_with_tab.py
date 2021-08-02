'''
this code shows how you can create different tabs with excel
basically, it is using the command:
workbook.create_sheet("nameOfTheWorkSheet")

'''

import openpyxl
# create the excel files with sheets
workbook = openpyxl.Workbook()

#create new sheet for each requirement term
reqTerm = ["1610", "1710", "1810", "1910", "2010"]
for rt in reqTerm:
    sheet = workbook.create_sheet(rt)
    sheet['A1'].value = "ReqTerm"
    sheet['B1'].value = "StudentID"
    sheet['C1'].value = "Name"
    sheet['D1'].value = "Handphone"


print(workbook.sheetnames)

workbook.save("excel_tabs.xlsx")


